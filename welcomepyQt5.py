import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QAction, QMenu, QVBoxLayout, QWidget, QPushButton, QDialog, QLineEdit, QInputDialog, QComboBox, QLabel, QListWidget, QMessageBox, QLabel, QDialogButtonBox
from PyQt5.QtCore import QDateTime, Qt, QTimer
from PyQt5.QtGui import QFont, QPixmap
import json
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import pandas as pd
import os
import getpass
import openpyxl
from PyQt5.QtGui import QColor
import webbrowser
import requests
import subprocess
from typing import NoReturn




# Fonts color
RESET = "\033[0m"             # Reset all formatting
BLACK = "\033[30m"            # Black text
RED = "\033[31m"              # Red text
GREEN = "\033[32m"            # Green text
YELLOW = "\033[33m"           # Yellow text
BLUE = "\033[34m"             # Blue text
MAGENTA = "\033[35m"          # Magenta text
CYAN = "\033[36m"             # Cyan text

text_color_red = QColor(255, 0, 0)  # Red color
text_color_green = QColor(159, 229, 124)  # Green color


now2 = time.strftime("%m")
today = time.strftime("%d/%m/%Y")
now1 = time.strftime("%d")


#Welcome
class WelcomeWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Welcome to my program')
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        welcome_label = QLabel('Welcome to Timekeeper!')
        welcome_label.setStyleSheet('font-size: 54px; font-weight: bold;')
        welcome_label.setFont(QFont('Times New Roman', 54, QFont.Bold))  # Set the font for the welcome label
        welcome_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(welcome_label)

        pixmap = QPixmap('timekeeper.jpg')  # Replace 'path_to_your_image.png' with the actual path to your image file

        welcome_label = QLabel()

        pixmap_resized = pixmap.scaled(1000, 800, Qt.AspectRatioMode.KeepAspectRatio)  # Resize the image to 400x300 pixels
        welcome_label.setPixmap(pixmap_resized)
        welcome_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(welcome_label)

        info_label = QLabel('This is my program. If it has any error, please contact me.')
        info_label.setFont(QFont('Times New Roman', 14))  # Set the font for the info label
        info_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(info_label)

        self.setLayout(layout)

        # Close the window after 5 seconds
        QTimer.singleShot(5000, self.close)

# Main
class MenuExample(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.initUI_1()

    def initUI(self):
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('Menu')

        # Timekeeper option
        timekeeper_action = QAction('Timekeeper', self)
        fileMenu.addAction(timekeeper_action)

        # Add Employee option
        add_employee_action = QAction('Add Employee', self)
        fileMenu.addAction(add_employee_action)

        # Delete Employee option
        delete_employee_action = QAction('Delete Employee', self)
        fileMenu.addAction(delete_employee_action)

        # Calculate Employee Salary option
        calculate_salary_action = QAction('Calculate Employee Salary', self)
        fileMenu.addAction(calculate_salary_action)

        # Pre-pay option
        pre_pay_action = QAction('Pre-pay', self)
        fileMenu.addAction(pre_pay_action)

        # Calculate Employee Salary after Pre-pay option
        calculate_salary_after_pay_action = QAction('Calculate Employee Salary after Pre-pay', self)
        fileMenu.addAction(calculate_salary_after_pay_action)

        # Export the Excel file option
        export_excel_action = QAction('Export the Excel file', self)
        fileMenu.addAction(export_excel_action)

        # Exit the Program option
        exit_action = QAction('Exit the Program', self)
        exit_action.triggered.connect(self.close)
        fileMenu.addAction(exit_action)

        # Create a widget and a layout for the buttons
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Create buttons and connect their clicked signals to respective functions
        timekeeper_button = QPushButton('Timekeeper')
        timekeeper_button.clicked.connect(self.timekeeper)
        timekeeper_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(timekeeper_button)

        add_employee_button = QPushButton('Add Employee')
        add_employee_button.clicked.connect(self.add_employee)
        add_employee_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(add_employee_button)

        delete_employee_button = QPushButton('Delete Employee')
        delete_employee_button.clicked.connect(self.delete_employee)
        delete_employee_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(delete_employee_button)

        calculate_salary_button = QPushButton('Calculate Employee Salary')
        calculate_salary_button.clicked.connect(self.calculate_salary)
        calculate_salary_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(calculate_salary_button)

        pre_pay_button = QPushButton('Pre-pay')
        pre_pay_button.clicked.connect(self.pre_pay)
        pre_pay_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(pre_pay_button)

        calculate_salary_after_pay_button = QPushButton('Calculate Employee Salary after Pre-pay')
        calculate_salary_after_pay_button.clicked.connect(self.calculate_salary_after_pay)
        calculate_salary_after_pay_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(calculate_salary_after_pay_button)

        export_excel_button = QPushButton('Export the Excel file')
        export_excel_button.clicked.connect(self.export_excel)
        export_excel_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(export_excel_button)

        exit_button = QPushButton('Exit the Program')
        exit_button.clicked.connect(self.exit)
        exit_button.setFont(QFont('Times New Roman', 14))  # Set the font for the button
        layout.addWidget(exit_button)

        self.setCentralWidget(widget)
        # self.setGeometry(500, 500, 500, 400)
        self.setWindowTitle('Timekeeper Menu')
        self.show()

        self.adjustSize()

    def initUI_1(self):
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('Help')

        # About us option
        about_action = QAction('About us', self)
        about_action.triggered.connect(self.about)  # Connect the action to the about slot
        fileMenu.addAction(about_action)

        # The Dev Profile option
        devprofile_action = QAction('My Profile', self)
        devprofile_action.triggered.connect(self.devprofile)  # Connect the action to the about slot
        fileMenu.addAction(devprofile_action)

        # About program option
        about_program_action = QAction('About the Program', self)
        about_program_action.triggered.connect(self.about_program)  # Connect the action to the about slot
        fileMenu.addAction(about_program_action)

        # About program option
        update_action = QAction('Update', self)
        update_action.triggered.connect(self.update)  # Connect the action to the about slot
        fileMenu.addAction(update_action)

    def about(self):
        # Assuming you have defined a color for the text
        text_color = QColor(0, 0, 0)  # Black color

        # Construct the HTML-formatted message with the desired color
        message_text = f"<font color='{text_color.name()}' size='4'>Hi! Thank you for using my Timekeeper program.</font><br>"
        message_text_1 = f"<font color='{text_color.name()}' size='4'>I hope you like it.</font><br>"
        message_text_2 = f"<font color='{text_color.name()}' size='4'>This was made by MrHecker.</font><br>"
        message_text_3 = f"<font color='{text_color.name()}' size='4'>Copyright 2023.</font>"

        message_box = QMessageBox()
        font = QFont("Times New Roman", 14)  # Specify the font family and size
        message_box.setFont(font)
        message_box.setWindowTitle("About us")


        # Construct the final message by concatenating the HTML-formatted texts
        message = "\n".join([message_text, message_text_1, message_text_2, message_text_3])

        # Set the HTML-formatted text as the informative text of the QMessageBox
        message_box.setInformativeText(message)

        # Show the QMessageBox
        message_box.exec()


    def devprofile(self):
        webbrowser.open("https://stackoverflow.com/users/18360467/fuckbkav1")
        webbrowser.open("https://github.com/7777Hecker")

    def about_program(self):
        
        text_color = QColor(0, 0, 0)

        message_text = f"""
            <font color="{text_color.name()}">
            You are in version 1.1</font><br>
            </font><br>
        """

        message_text_2 = f"""
            <font color="{text_color.name()}">
            What new in this version:
            </font><br>
        """

        message_text_3 = f"""
            <font color="{text_color.name()}">
            Fix the size of yes_no_dialog
            <br></br> Add more options to choose from (like Yes or No or Exit)
            <br></br>Update the code that timekeeper normal or khoang</font><br>
            </font><br>
        """

        message_text_1 = f"""
            <font color="{text_color.name()}">
            If the new version has please press the 'Update' button to download the newest version from me. 
            <br></br>That may have you more options. Thanks!
            <br></br> Can't use now because it in fix and update. Sorry for this inconvenient.<br></br>
            </font><br>
        """

        message_box = QMessageBox()
        font = QFont("Times New Roman", 14)  # Specify the font family and size
        message_box.setFont(font)
        message_box.setWindowTitle("About the program")


        # Construct the final message by concatenating the HTML-formatted texts
        message = "\n".join([message_text, message_text_2, message_text_3, message_text_1])

        # Set the HTML-formatted text as the informative text of the QMessageBox
        message_box.setInformativeText(message)

        # Show the QMessageBox
        message_box.exec()


    def update(self):
        LATEST_VERSION = "1.0" 
        def check_update():
            r = requests.get("https://example.com/latest-version.txt")
            latest_version = r.text
            if latest_version > LATEST_VERSION:
                return True
            else: 
                return False
            
        def show_update_dialog():

            dialog = QDialog()
            dialog.setWindowTitle('Has New Version')

            layout = QVBoxLayout()

            message = QLabel('The newest version has been had, do you want to update it?')  
            layout.addWidget(message)

            button_box = QDialogButtonBox(QDialogButtonBox.Yes | QDialogButtonBox.No)

            button_box.accepted.connect(dialog.accept)  
            button_box.rejected.connect(dialog.reject)

            layout.addWidget(button_box)  
            dialog.setLayout(layout)

            reply = dialog.exec_()

            return reply
        
        def show_success_message():

            msg = QMessageBox()
            msg.setWindowTitle("Update success")
            msg.setText("The newest version has been isntalled.\nThe program will automatic close down. Sorry for this inconvenient")
            msg.setIcon(QMessageBox.Information)

            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

        def download_file(url, file_path):
            response = requests.get(url, stream=True)
            if response.ok:
                with open(file_path, 'wb') as f:  
                    for chunk in response.iter_content(chunk_size=1024):
                        if chunk:
                            f.write(chunk)
            else:
                print("Download failed")
                return
            print("File downloaded")
            show_success_message()
            exit()
        
        def download_update_1():
            url = "https://example.com/file.py"
            path = "new_file.py"
            
            download_file(url, path)
            
            # Notification that download success
            show_success_message()  

            # reboot the program  
            exit()

        def run():
            if check_update():
                reply = show_update_dialog()
                if reply == QDialog.Accepted:
                    download_update_1()

    def timekeeper(self):
        class TimekeeperWindow(QDialog):
            def __init__(self):
                super().__init__()
                self.setWindowTitle("Timekeeper")
                self.setGeometry(0, 0, 396, 974)

                self.employee_data = self.load_employee_data()
                self.employee_data_half = self.load_employee_data_half()

                self.label = QLabel("Select an employee:")
                self.button_layout = QVBoxLayout()

                for employee_id in self.employee_data.keys():
                    button = QPushButton(employee_id)
                    button.setFont(QFont("Times New Roman", 14))  # Specify the font family and size for the button
                    button.clicked.connect(self.timekeep_employee)
                    self.button_layout.addWidget(button)

                central_widget = QWidget()
                central_widget.setLayout(self.button_layout)
                self.setLayout(self.button_layout)

            def load_employee_data(self):
                with open("employee_data.json", "r") as file:
                    employee_data = json.load(file)
                return employee_data

            def load_employee_data_half(self):
                with open("employee_data_half.json", "r") as file:
                    employee_data_half = json.load(file)
                return employee_data_half

            def save_employee_data(self):
                with open("employee_data.json", "w") as file:
                    json.dump(self.employee_data, file)

            def save_employee_data_half(self):
                with open("employee_data_half.json", "w") as file:
                    json.dump(self.employee_data_half, file)

            def timekeep_employee(self):
                employee_id = self.sender().text()

                current_time = QDateTime.currentDateTime()
                timestamp = current_time.toString("dd/MM/yyyy HH:mm:ss")

                if employee_id in self.employee_data:
                    self.employee_data[employee_id]["timekeeping"].append(timestamp)

                    selected_option = self.yes_no_dialog_khoang()
                    khoang = selected_option


                    if khoang == 'No':

                        selected_option = self.yes_no_dialog()
                        half_day = selected_option

                        if half_day == "Yes":
                            if employee_id not in self.employee_data_half:
                                self.employee_data_half[employee_id] = {}
                            if "timekeeping_half" not in self.employee_data_half[employee_id]:
                                self.employee_data_half[employee_id]["timekeeping_half"] = []
                            self.employee_data_half[employee_id]["timekeeping_half"].append(timestamp)

                        if half_day == "No":
                            self.save_employee_data()
                        elif half_day == "Yes":
                            self.save_employee_data_half()
                        elif half_day == "Exit":
                            pass

                        if half_day == "Yes" or half_day == 'No':
                            print(GREEN + "Timekeep succeed:", self.employee_data[employee_id]["name"] + RESET)
                            with open("employee_timekeep.txt", "a+", encoding="utf-8") as abc:
                                abc.write("\n" + "Employee: " + str(self.employee_data[employee_id]["name"]) + " at " + timestamp)

                        if half_day == "Yes":
                            data = {'Ngày': QDateTime.currentDateTime().toString("dd/MM/yyyy"), 'Tên': employee_id, 'Làm việc': 'X/2'}
                        elif half_day == 'No':
                            data = {'Ngày': QDateTime.currentDateTime().toString("dd/MM/yyyy"), 'Tên': employee_id, 'Làm việc': 'X'}
                        else:
                            pass
                        
                        if half_day == "Yes" or half_day == 'No':
                            df = pd.DataFrame(data, index=[0])

                            try:
                                existing_data = pd.read_excel('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                                df = pd.concat([existing_data, df], ignore_index=True)
                            except FileNotFoundError:
                                pass
                    elif khoang == 'Yes':

                        selected_option = self.yes_no_dialog()
                        half_day = selected_option

                        if half_day == "Yes":
                            if employee_id not in self.employee_data_half:
                                self.employee_data_half[employee_id] = {}
                            if "timekeeping_half" not in self.employee_data_half[employee_id]:
                                self.employee_data_half[employee_id]["timekeeping_half"] = []
                            self.employee_data_half[employee_id]["timekeeping_half"].append(timestamp)

                        if half_day == "No":
                            self.save_employee_data()
                        elif half_day == "Yes":
                            self.save_employee_data_half()
                        elif half_day == "Exit":
                            pass

                        if half_day == "Yes" or half_day == 'No':
                            print(GREEN + "Timekeep succeed:", self.employee_data[employee_id]["name"] + RESET)
                            with open("employee_timekeep.txt", "a+", encoding="utf-8") as abc:
                                abc.write("\n" + "Employee: " + str(self.employee_data[employee_id]["name"]) + " at " + timestamp)

                        if half_day == "Yes":
                            data = {'Ngày': QDateTime.currentDateTime().toString("dd/MM/yyyy"), 'Tên': employee_id, 'Làm việc': 'X/2', 'Khoáng': khoang}
                        elif half_day == 'No':
                            data = {'Ngày': QDateTime.currentDateTime().toString("dd/MM/yyyy"), 'Tên': employee_id, 'Làm việc': 'X', 'Khoáng': khoang}
                        else:
                            pass
                        
                        if half_day == "Yes" or half_day == 'No':
                            df = pd.DataFrame(data, index=[0])

                            try:
                                existing_data = pd.read_excel('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                                df = pd.concat([existing_data, df], ignore_index=True)
                            except FileNotFoundError:
                                pass
                    
                    elif khoang == 'Exit':
                        pass

                    if khoang == 'Yes' and half_day == 'No':
                        df.to_excel('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx', index=False, sheet_name='Chấm công')

                        workbook = load_workbook('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                        sheet = workbook.active
                        font = Font(size=14, bold=False, name='Times New Roman')
                        font1 = Font(size=14, bold=True, name='Times New Roman')

                        thin_border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))

                        for column_letter in ['A', 'B', 'C', 'D']:
                            for cell in sheet[column_letter][1:]:
                                cell.font = font
                                cell.border = thin_border
                                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                sheet.column_dimensions[column_letter].width = column_width + 10

                        sheet['A1'].font = font1
                        sheet['B1'].font = font1
                        sheet['C1'].font = font1
                        sheet['D1'].font = font1

                        workbook.save('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                    elif khoang == 'Yes' and half_day == 'Yes':
                        df.to_excel('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx', index=False, sheet_name='Chấm công')
                        workbook = load_workbook('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                        sheet = workbook.active
                        font = Font(size=14, bold=False, name='Times New Roman')
                        font1 = Font(size=14, bold=True, name='Times New Roman')

                        thin_border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))

                        for column_letter in ['A', 'B', 'C', 'D']:
                            for cell in sheet[column_letter][1:]:
                                cell.font = font
                                cell.border = thin_border
                                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                sheet.column_dimensions[column_letter].width = column_width + 10

                        sheet['A1'].font = font1
                        sheet['B1'].font = font1
                        sheet['C1'].font = font1
                        sheet['D1'].font = font1

                        workbook.save('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                    elif khoang == 'No' and half_day == 'Yes':
                        df.to_excel('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx', index=False, sheet_name='Chấm công')
                        workbook = load_workbook('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                        sheet = workbook.active
                        font = Font(size=14, bold=False, name='Times New Roman')
                        font1 = Font(size=14, bold=True, name='Times New Roman')

                        thin_border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))

                        for column_letter in ['A', 'B', 'C', 'D']:
                            for cell in sheet[column_letter][1:]:
                                cell.font = font
                                cell.border = thin_border
                                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                sheet.column_dimensions[column_letter].width = column_width + 10

                        sheet['A1'].font = font1
                        sheet['B1'].font = font1
                        sheet['C1'].font = font1
                        sheet['D1'].font = font1

                        workbook.save('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                    elif khoang == 'No' and half_day == 'No':
                        df.to_excel('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx', index=False, sheet_name='Chấm công')
                        workbook = load_workbook('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                        sheet = workbook.active
                        font = Font(size=14, bold=False, name='Times New Roman')
                        font1 = Font(size=14, bold=True, name='Times New Roman')

                        thin_border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))

                        for column_letter in ['A', 'B', 'C', 'D']:
                            for cell in sheet[column_letter][1:]:
                                cell.font = font
                                cell.border = thin_border
                                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                sheet.column_dimensions[column_letter].width = column_width + 10

                        sheet['A1'].font = font1
                        sheet['B1'].font = font1
                        sheet['C1'].font = font1
                        sheet['D1'].font = font1

                        workbook.save('cham_cong_thang ' + QDateTime.currentDateTime().toString(str(now2)) + '.xlsx')
                    elif khoang == 'Exit' and half_day == 'Exit':
                        pass
                    
                    

            def yes_no_dialog(self):
                buttons = ("Yes", "No", "Exit")
                options = []
                for button_text in buttons:
                    button = QPushButton(button_text)
                    button.setFont(QFont("Times New Roman", 14))
                    button.clicked.connect(self.dialog_button_clicked)
                    options.append(button)

                dialog = QDialog(self)
                dialog.setWindowModality(Qt.ApplicationModal)
                dialog_layout = QVBoxLayout()

                for option in options:
                    dialog_layout.addWidget(option)

                dialog_layout.setSizeConstraint(QVBoxLayout.SetFixedSize)

                dialog.setLayout(dialog_layout)
                
                # Set a custom font for the WindowTitle (DialogTitle)
                title_label = QLabel("Did the employee half a day?")
                title_label.setFont(QFont("Times New Roman", 16))
                dialog.setWindowTitle(" ")  # Set a space to avoid the default title
                dialog.setWindowFlags(Qt.Dialog | Qt.CustomizeWindowHint)
                
                # Add the title label to the layout
                dialog_layout.insertWidget(0, title_label)

                dialog.exec_()

                return self.selected_option
            def yes_no_dialog_khoang(self):
                buttons = ("Yes", "No", "Exit")
                options = []
                for button_text in buttons:
                    button = QPushButton(button_text)
                    button.setFont(QFont("Times New Roman", 14))
                    button.clicked.connect(self.dialog_button_clicked)
                    options.append(button)

                dialog = QDialog(self)
                dialog.setWindowModality(Qt.ApplicationModal)
                dialog_layout = QVBoxLayout()

                for option in options:
                    dialog_layout.addWidget(option)

                dialog_layout.setSizeConstraint(QVBoxLayout.SetFixedSize)

                dialog.setLayout(dialog_layout)
                
                # Set a custom font for the WindowTitle (DialogTitle)
                title_label = QLabel("Khoang?")
                title_label.setFont(QFont("Times New Roman", 16))
                dialog.setWindowTitle(" ")  # Set a space to avoid the default title
                dialog.setWindowFlags(Qt.Dialog | Qt.CustomizeWindowHint)
                
                # Add the title label to the layout
                dialog_layout.insertWidget(0, title_label)

                dialog.exec_()

                return self.selected_option
            def dialog_button_clicked(self):
                self.selected_option = self.sender().text()
                dialog = self.sender().parent()
                dialog.close()

        window = TimekeeperWindow()
        window.exec_()

    def add_employee(self):
        def load_employee_data():
            try:
                with open("employee_data.json", "r", encoding="utf-8") as file:
                    return json.load(file)
            except FileNotFoundError:
                return {}

        def save_employee_data(data):
            with open("employee_data.json", "w", encoding="utf-8") as file:
                json.dump(data, file)

        def save_employee_data_backup(employee_data):
            with open("employee_data_backup.txt", "w") as file:
                file.write(str(employee_data))

        
        class AddEmployeeDialog(QDialog):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Add Employee")

                self.label_employee = QLabel("Input the employee:")
                self.line_edit_employee = QLineEdit()
                self.label_employee.setFont(QFont("Times New Roman", 14))  # Set the font for the label
                self.line_edit_employee.setFont(QFont("Times New Roman", 14))  # Set the font for the line edit

                self.label_salary = QLabel("The salary of employee per day:")
                self.line_edit_salary = QLineEdit()
                self.label_salary.setFont(QFont("Times New Roman", 14))  # Set the font for the label
                self.line_edit_salary.setFont(QFont("Times New Roman", 14))  # Set the font for the line edit

                self.button_add = QPushButton("Add")
                self.button_add.setFont(QFont("Times New Roman", 14))  # Set the font for the button
                self.button_add.clicked.connect(self.add_employee)

                layout = QVBoxLayout()
                layout.addWidget(self.label_employee)
                layout.addWidget(self.line_edit_employee)
                layout.addWidget(self.label_salary)
                layout.addWidget(self.line_edit_salary)
                layout.addWidget(self.button_add)

                self.setLayout(layout)

            def add_employee(self):
                employee_id = self.line_edit_employee.text()
                input_text = self.line_edit_salary.text()
                if input_text.strip():  # Check if the input is not empty after removing leading/trailing whitespace
                    date_rate = float(input_text)
                else:
                    date_rate = 0
                employee_data = load_employee_data()
                if employee_id in employee_data:
                    print(RED + "The employee already exists." + RESET)
                else:
                    employee_data[employee_id] = {"name": employee_id, "data_rate": date_rate, "timekeeping": []}
                    save_employee_data(employee_data)
                    save_employee_data_backup(employee_data)
                    print(GREEN + "Add employee succeed" + RESET, employee_id)
                    with open("employee.txt", 'a+', encoding='utf-8') as a45:
                        a45.write("\n" + str(employee_id))

                self.close()

        def add_employee():
            dialog = AddEmployeeDialog()
            dialog.exec_()

        add_employee()

    def delete_employee(self):
        class DeleteEmployeeWindow(QDialog):
            def __init__(self):
                super().__init__()
                self.setWindowTitle('Delete Employee')
                self.setGeometry(350, 250, 580, 450)

                self.employee_data = self.load_employee_data()
                self.employee_list = list(self.employee_data.keys())

                self.label = QLabel('Select the employee to delete:', self)
                self.label.setGeometry(50, 70, 420, 60)
                self.label.setFont(QFont("Times New Roman", 14))  # Set the font for the label


                self.employee_combobox = QComboBox(self)
                self.employee_combobox.setGeometry(100, 120, 210, 50)
                self.populate_employee_combobox()

                self.delete_button = QPushButton('Delete', self)
                self.delete_button.setGeometry(350, 300, 200, 90)
                self.delete_button.clicked.connect(self.delete_employee)
                self.delete_button.setFont(QFont("Times New Roman", 14))  # Set the font for the button


            def load_employee_data(self):
                try:
                    with open("employee_data.json", "r") as file:
                        return json.load(file)
                except FileNotFoundError:
                    return {}

            def save_employee_data(self):
                with open("employee_data.json", "w") as file:
                    json.dump(self.employee_data, file)

            

            def populate_employee_combobox(self):
                self.employee_combobox.addItems(self.employee_list)

            def delete_employee(self):
                def delete_digit_in_file(file_path, digit):
                    # Read the file
                    with open(file_path, 'r') as file:
                        content = file.read()

                    # Remove the digit from the content
                    content = content.replace(digit, '')

                    # Write the updated content back to the file
                    with open(file_path, 'w') as file:
                        file.write(content)

                employee_id = self.employee_combobox.currentText()
                if employee_id in self.employee_data:
                    del self.employee_data[employee_id]
                    self.save_employee_data()
                    self.employee_list.remove(employee_id)
                    self.employee_combobox.removeItem(self.employee_combobox.currentIndex())
                    message_box = QMessageBox()
                    message_box.setIcon(QMessageBox.Information)
                    message_box.setText(f"{employee_id} has been deleted.")
                    message_box.setWindowTitle("Success")
                    message_box.setFont(QFont("Times New Roman", 14))  # Set the font for the message box
                    file_path = 'employee.txt'  # Replace with your file path
                    digit_to_delete = employee_id      # Replace with the digit you want to delete
                    delete_digit_in_file(file_path, digit_to_delete)
                    message_box.exec_()
                    print(GREEN + "Delete employee succeed" + RESET, employee_id)
                else:
                    message_box = QMessageBox()
                    message_box.setIcon(QMessageBox.Warning)
                    message_box.setText("Invalid employee selected.")
                    message_box.setWindowTitle("Error")
                    message_box.setFont(QFont("Times New Roman", 14))  # Set the font for the message box
                    message_box.exec_()
                    print(RED + "Invalid employee selected" + RESET)

        delete_employee_window = DeleteEmployeeWindow()
        delete_employee_window.exec_()


    def calculate_salary(self):

        def load_employee_data():
            with open("employee_data.json", "r") as file:
                employee_data = json.load(file)
            return employee_data

        # Read the "employee_data_pre_pay.json" from the system
        def load_employee_data1():
            try:
                with open("employee_data_pre_pay.json", "r") as file:
                    employee_data = eval(file.read())
                    return employee_data
            except FileNotFoundError:
                return {}

        # Read the "employee_data_half.json" from the system
        def load_employee_data_half():
            with open("employee_data_half.json", "r") as file:
                employee_data_half = json.load(file)
            return employee_data_half

        # Save the "employee_data_half.json" to the system
        def save_employee_data_half(employee_data_half):
            with open("employee_data_half.json", "w") as file:
                json.dump(employee_data_half, file)

        # Save the "employee_data.txt" to the system
        def save_employee_data(employee_data):
            with open("employee_data.txt", "w") as file:
                file.write(str(employee_data))

        ac = ''
        employee_data = load_employee_data()
        employee_data_half = load_employee_data_half()
        logfile = open('employee.txt','r',encoding='utf-8')
        f = logfile.read()
        b = f.split()
        today = time.strftime("%d/%m/%Y")
        employee_name = []
        i2 = 0

        workbook_path = 'cham_cong_luong_thang ' + str(now2) + '.xlsx'

        try:
            existing_data = pd.read_excel(workbook_path)
        except FileNotFoundError:
            existing_data = pd.DataFrame()

        data = {'STT': [], 'Tên': [], 'Lương Trước Khi Ứng': [], 'Số Công': []}

        for i in range(0, len(b)):
            print("")
            ac = ac + str(i)
            find1 = ac.find(str(i))
            if 0 <= find1 < len(b):
                employee_id = b[find1]
        
                if employee_id in employee_data:
                    employee = employee_data[employee_id]

                    if employee_id in employee_data_half:
                        employee_half = employee_data_half[employee_id]
                        work_hours = len(employee["timekeeping"])
                        work_hours_half = len(employee_data_half.get(employee_id, {}).get("timekeeping_half", []))
                        salary_half = work_hours_half * (employee["data_rate"] / 2)
                        salary_full = work_hours * employee["data_rate"]
                        salary = salary_half + salary_full
                        timekeeper_employee = work_hours + (work_hours_half/2)
                    else:
                        work_hours = len(employee["timekeeping"])
                        salary_full = work_hours * employee["data_rate"]
                        salary = salary_full
                        timekeeper_employee = work_hours

                    employee_name.append(f"Employee salary {employee['name']} is: {salary}. ")
                    with open("employee_salary.txt","a+",encoding="utf-8") as a4:
                        a4.write("\n" + employee["name"] + ": " + str(salary))

                    data['Tên'].append(employee_id)
                    data['Lương Trước Khi Ứng'].append(salary)
                    data['Số Công'].append(timekeeper_employee)
                    i2 = i2 + 1
                    data['STT'].append(str(i2))

                    message_text = employee_name[i]
                    message = "\n".join([message_text])

                else:
                    employee_name.append(f"Can't see the employee:  {employee_id}. ")

            df = pd.DataFrame(data)
            df = pd.concat([existing_data, df], ignore_index=True)
            df.to_excel(workbook_path, index=False, sheet_name='Lương')

            # Load the existing Excel file
            workbook = load_workbook(workbook_path)

            # Get the active sheet
            sheet = workbook.active

            # Set the font size and format for the entire column A
            font = Font(size=14, bold=False, name='Times New Roman')
            font1 = Font(size=14, bold=True, name='Times New Roman')

            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

            sheet['A1'].font = font1
            sheet['B1'].font = font1
            sheet['C1'].font = font1
            sheet['D1'].font = font1
            sheet['E1'].font = font1

            column_letter = 'A'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            column_letter = 'B'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            column_letter = 'C'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10
            
            column_letter = 'D'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10
            
            column_letter = 'E'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            workbook.save(workbook_path)

            message_box = QMessageBox()
            font = QFont("Times New Roman", 14)  # Specify the font family and size
            message_box.setFont(font)
            message_box.setWindowTitle("Salary")
    
            # Set the HTML-formatted text as the informative text of the QMessageBox
            message_box.setInformativeText(message)

            # Show the QMessageBox
            message_box.exec()


    def pre_pay(self):
        class PrePayWindow(QDialog):
            def load_employee_data1(self):
                try:
                    with open("employee_data_pre_pay.json", "r") as file:
                        employee_data = eval(file.read())
                        return employee_data
                except FileNotFoundError:
                    return {}
                
            def load_employee_data2(self):
                try:
                    with open("employee.txt", "r") as file:
                        employee_data = eval(file.read())
                        return employee_data
                except FileNotFoundError:
                    return {}

            def load_employee_data(self):
                with open("employee_data.json", "r") as file:
                    employee_data = json.load(file)
                return employee_data

            def load_employee_data_half(self):
                with open("employee_data_half.json", "r") as file:
                    employee_data_half = json.load(file)
                return employee_data_half

            def save_employee_data(self):
                with open("employee_data.json", "w") as file:
                    json.dump(self.employee_data, file)

            def save_employee_data_half(self):
                with open("employee_data_half.json", "w") as file:
                    json.dump(self.employee_data_half, file)

            def __init__(self):
                super().__init__()
                self.setWindowTitle('Pre-Pay')
                self.setGeometry(100, 100, 400, 300)

                self.employee_data = self.load_employee_data1()
                self.employee_data1 = self.load_employee_data()
                self.employee_data_half = self.load_employee_data_half()

                self.label = QLabel('Select the employee to pre-pay:', self)
                self.label.setGeometry(50, 50, 300, 30)
                self.label.setFont(QFont("Times New Roman", 14))  # Set the font for the label


                self.employee_combobox = QComboBox(self)
                self.employee_combobox.setGeometry(100, 90, 210, 50)
                self.populate_employee_combobox()

                self.pre_pay_button = QPushButton('Selected', self)
                self.pre_pay_button.setGeometry(150, 180, 200, 90)
                self.pre_pay_button.setFont(QFont("Times New Roman", 14))  # Set the font for the button
                self.pre_pay_button.clicked.connect(self.pre_pay)

            def populate_employee_combobox(self):
                for employee_id in self.employee_data1:
                    self.employee_combobox.addItem(employee_id)

            def pre_pay(self):
                employee_id = self.employee_combobox.currentText()

                current_time = QDateTime.currentDateTime().toString('dd/MM/yyyy HH:mm:ss')

                if employee_id in self.employee_data:
                    employee_pre_pay = self.employee_data[employee_id]
                    if "data_rate_1" in employee_pre_pay:
                        font = QFont("Times New Roman", 14)  # Specify the font family and size
                        dialog = QDialog(self)
                        dialog.setWindowTitle("Pre-Pay")
                        dialog.setModal(False)  # Allow user interaction with the main window

                        dialog_layout = QVBoxLayout()

                        label = QLabel("Enter the amount to pre-pay:")
                        label.setFont(font)
                        dialog_layout.addWidget(label)

                        line_edit = QLineEdit()
                        line_edit.setFont(font)
                        dialog_layout.addWidget(line_edit)

                        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
                        dialog_layout.addWidget(button_box)

                        dialog.setLayout(dialog_layout)

                        def handle_dialog_result(result):
                            if result == QDialogButtonBox.Ok:
                                money_text = line_edit.text()
                                if money_text:
                                    money = float(money_text)
                                    money1 = float(money_text)
                                    employee_pre_pay = self.employee_data[employee_id]
                                    salary_pre_pay = employee_pre_pay["data_rate_1"]
                                    money = float(money) + float(salary_pre_pay)

                                    print(GREEN + "Pre-payment succeeded for employee", employee_id, "at:", current_time + RESET)

                                    with open("employee_pre_pay.txt", "a+", encoding="utf-8") as abc:
                                        abc.write("\nPre_pay: " + str(money) + " succeeded for " + str(employee_id) + " at: " + current_time)

                                    if employee_id in self.employee_data1:
                                        self.employee_data[employee_id]["data_rate_1"] = money
                                        with open("employee_data_pre_pay.json", "w") as file:
                                            json.dump(self.employee_data, file)
                                        print(GREEN + "employee_data_pre_pay:", money, "succeeded for", employee_id, "at:", current_time + RESET)

                                        data = {'Ngày': today, 'Tên': employee_id, 'Tiền ứng': money1}
                                        df = pd.DataFrame(data, index=[0])
                                        try:
                                            existing_data = pd.read_excel('cham_cong_ung_thang ' + str(now2) + '.xlsx')
                                            df = pd.concat([existing_data, df], ignore_index=True)
                                        except FileNotFoundError:
                                            pass
                                        df.to_excel('cham_cong_ung_thang ' + str(now2) + '.xlsx', index=False, sheet_name='Tiền Ứng')

                                        workbook = load_workbook('cham_cong_ung_thang ' + str(now2) + '.xlsx')
                                        sheet = workbook.active

                                        font = Font(size=14, bold=False, name='Times New Roman')
                                        font1 = Font(size=14, bold=True, name='Times New Roman')
                                        sheet['A1'].font = font1
                                        sheet['B1'].font = font1
                                        sheet['C1'].font = font1

                                        thin_border = Border(left=Side(style='thin'), 
                                                            right=Side(style='thin'), 
                                                            top=Side(style='thin'), 
                                                            bottom=Side(style='thin'))

                                        column_letter = 'A'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        column_letter = 'B'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        column_letter = 'C'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        column_letter = 'D'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        workbook.save('cham_cong_ung_thang ' + str(now2) + '.xlsx')
                                    else:
                                        print(RED + "Can't find the employee", employee_id + RESET)

                        button_box.accepted.connect(lambda: handle_dialog_result(QDialogButtonBox.Ok))

                        dialog.exec_()

                    else:
                        dialog = QDialog(self)
                        font = QFont("Times New Roman", 14)  # Specify the font family and size
                        dialog.setWindowTitle("Pre-Pay")
                        dialog.setModal(False)  # Allow user interaction with the main window

                        dialog_layout = QVBoxLayout()

                        label = QLabel("Enter the amount to pre-pay:")
                        label.setFont(font)
                        dialog_layout.addWidget(label)

                        line_edit = QLineEdit()
                        line_edit.setFont(font)
                        dialog_layout.addWidget(line_edit)

                        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
                        dialog_layout.addWidget(button_box)

                        dialog.setLayout(dialog_layout)
                        
                        def handle_dialog_result(result):
                            if result == QDialogButtonBox.Ok:
                                money_text = line_edit.text()
                                if money_text:
                                    money = float(money_text)
                                    money1 = float(money_text)

                                    print(GREEN + "Pre-payment succeeded for employee", employee_id, "at:", current_time + RESET)

                                    with open("employee_pre_pay.txt", "a+", encoding="utf-8") as abc:
                                        abc.write("\nPre_pay: " + str(money) + " succeeded for " + str(employee_id) + " at: " + current_time)

                                    if employee_id in self.employee_data1:
                                        self.employee_data[employee_id] = {"name": employee_id, "data_rate_1": money, "pre_pay": []}
                                        with open("employee_data_pre_pay.json", "w") as file:
                                            json.dump(self.employee_data, file)
                                        print(GREEN + "Employee_data_pre_pay:", money, "succeeded for", employee_id, "at:", current_time + RESET)

                                        data = {'Ngày': today, 'Tên': employee_id, 'Tiền ứng': money}
                                        df = pd.DataFrame(data, index=[0])
                                        try:
                                            existing_data = pd.read_excel('cham_cong_ung_thang ' + str(now2) + '.xlsx')
                                            df = pd.concat([existing_data, df], ignore_index=True)
                                        except FileNotFoundError:
                                            pass
                                        df.to_excel('cham_cong_ung_thang ' + str(now2) + '.xlsx', index=False, sheet_name='Tiền Ứng')

                                        workbook = load_workbook('cham_cong_ung_thang ' + str(now2) + '.xlsx')
                                        sheet = workbook.active
                                        font = Font(size=14, bold=False, name='Times New Roman')
                                        font1 = Font(size=14, bold=True, name='Times New Roman')
                                        sheet['A1'].font = font1
                                        sheet['B1'].font = font1
                                        sheet['C1'].font = font1
                                        sheet['D1'].font = font1

                                        thin_border = Border(left=Side(style='thin'), 
                                                            right=Side(style='thin'), 
                                                            top=Side(style='thin'), 
                                                            bottom=Side(style='thin'))

                                        column_letter = 'A'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        column_letter = 'B'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        column_letter = 'C'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        column_letter = 'D'
                                        for cell in sheet[column_letter][1:]:
                                            cell.font = font
                                            cell.border = thin_border
                                            column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                            sheet.column_dimensions[column_letter].width = column_width + 10

                                        workbook.save('cham_cong_ung_thang ' + str(now2) + '.xlsx')

                                    else:
                                        print(RED + "Can't find the employee", employee_id + RESET)
                        
                        button_box.accepted.connect(lambda: handle_dialog_result(QDialogButtonBox.Ok))

                        dialog.exec_()

                else:
                    dialog = QDialog(self)
                    font = QFont("Times New Roman", 14)  # Specify the font family and size
                    dialog.setWindowTitle("Pre-Pay")
                    dialog.setModal(False)  # Allow user interaction with the main window
                    dialog_layout = QVBoxLayout()
                    label = QLabel("Enter the amount to pre-pay:")
                    label.setFont(font)
                    dialog_layout.addWidget(label)
                    line_edit = QLineEdit()
                    line_edit.setFont(font)
                    dialog_layout.addWidget(line_edit)

                    button_box = QDialogButtonBox(QDialogButtonBox.Ok)
                    dialog_layout.addWidget(button_box)

                    dialog.setLayout(dialog_layout)

                    def handle_dialog_result(result):
                        if result == QDialogButtonBox.Ok:
                            money_text = line_edit.text()
                            money_text = line_edit.text()
                            if money_text:
                                money = float(money_text)
                                money1 = float(money_text)
                                print(GREEN + "Pre-payment succeeded for employee", employee_id, "at:", current_time + RESET)
                                with open("employee_pre_pay.txt", "a+", encoding="utf-8") as abc:
                                    abc.write("\nPre_pay: " + str(money) + " succeeded for " + str(employee_id) + " at: " + current_time)
                                if employee_id in self.employee_data1:
                                    self.employee_data1[employee_id] = {"name": employee_id, "data_rate_1": money, "pre_pay": []}
                                    with open("employee_data_pre_pay.json", "w") as file:
                                        json.dump(self.employee_data1, file)
                                    print(GREEN + "Employee_data_pre_pay:", money, "succeeded for", employee_id, "at:", current_time + RESET)

                                    data = {'Ngày': today, 'Tên': employee_id, 'Tiền ứng': money}
                                    df = pd.DataFrame(data, index=[0])
                                    try:
                                        existing_data = pd.read_excel('cham_cong_ung_thang ' + str(now2) + '.xlsx')
                                        df = pd.concat([existing_data, df], ignore_index=True)
                                    except FileNotFoundError:
                                        pass
                                    df.to_excel('cham_cong_ung_thang ' + str(now2) + '.xlsx', index=False, sheet_name='Tiền Ứng')

                                    workbook = load_workbook('cham_cong_ung_thang ' + str(now2) + '.xlsx')
                                    sheet = workbook.active

                                    font = Font(size=14, bold=False, name='Times New Roman')
                                    font1 = Font(size=14, bold=True, name='Times New Roman')
                                    sheet['A1'].font = font1
                                    sheet['B1'].font = font1
                                    sheet['C1'].font = font1
                                    sheet['D1'].font = font1

                                    thin_border = Border(left=Side(style='thin'), 
                                                        right=Side(style='thin'), 
                                                        top=Side(style='thin'), 
                                                        bottom=Side(style='thin'))

                                    column_letter = 'A'
                                    for cell in sheet[column_letter][1:]:
                                        cell.font = font
                                        cell.border = thin_border
                                        column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                        sheet.column_dimensions[column_letter].width = column_width + 10

                                    column_letter = 'B'
                                    for cell in sheet[column_letter][1:]:
                                        cell.font = font
                                        cell.border = thin_border
                                        column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                        sheet.column_dimensions[column_letter].width = column_width + 10

                                    column_letter = 'C'
                                    for cell in sheet[column_letter][1:]:
                                        cell.font = font
                                        cell.border = thin_border
                                        column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                        sheet.column_dimensions[column_letter].width = column_width + 10

                                    column_letter = 'D'
                                    for cell in sheet[column_letter][1:]:
                                        cell.font = font
                                        cell.border = thin_border
                                        column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                                        sheet.column_dimensions[column_letter].width = column_width + 10

                                    workbook.save('cham_cong_ung_thang ' + str(now2) + '.xlsx')

                                else:
                                    print(RED + "Can't find the employee", employee_id + RESET)
                            
                    button_box.accepted.connect(lambda: handle_dialog_result(QDialogButtonBox.Ok))

                    dialog.exec_()

        prepay_window = PrePayWindow()
        prepay_window.exec_()

    def calculate_salary_after_pay(self):
        
        def load_employee_data():
            with open("employee_data.json", "r") as file:
                employee_data = json.load(file)
            return employee_data

        # Read the "employee_data_pre_pay.json" from the system
        def load_employee_data1():
            try:
                with open("employee_data_pre_pay.json", "r") as file:
                    employee_data = eval(file.read())
                    return employee_data
            except FileNotFoundError:
                return {}

        # Read the "employee_data_half.json" from the system
        def load_employee_data_half():
            with open("employee_data_half.json", "r") as file:
                employee_data_half = json.load(file)
            return employee_data_half

        # Save the "employee_data_half.json" to the system
        def save_employee_data_half(employee_data_half):
            with open("employee_data_half.json", "w") as file:
                json.dump(employee_data_half, file)

        # Save the "employee_data.txt" to the system
        def save_employee_data(employee_data):
            with open("employee_data.txt", "w") as file:
                file.write(str(employee_data))

        ac = ''
        employee_data = load_employee_data()  # Load file timekeeping
        employee_data1 = load_employee_data1()  # Load file pre-pay
        employee_data_half = load_employee_data_half()  # Load file timekeeping half
        logfile = open('employee.txt', 'r', encoding='utf-8')
        f = logfile.read()
        b = f.split()
        today = time.strftime("%d/%m/%Y")
        employee_ids = []
        salary_pre_pays = []
        employee_name = []
        timekeeper_employees = []
        i2 = 0
        i3 = []


        workbook_path = 'cham_cong_luong_thang ' + str(now2) + '.xlsx'

        try:
            existing_data = pd.read_excel(workbook_path)
        except FileNotFoundError:
            existing_data = pd.DataFrame()
        
        for i in range(0, len(b)):
            salary_pre_pay = 0
            print("")
            ac = ac + str(i)

            find1 = ac.find(str(i))
            if 0 <= find1 < len(b):
                employee_id = b[find1]
                if employee_id in employee_data:
                    employee = employee_data[employee_id]
                    if employee_id in employee_data_half:
                        employee_half = employee_data_half[employee_id]
                        work_hours = len(employee["timekeeping"])
                        work_hours_half = len(employee_data_half.get(employee_id, {}).get("timekeeping_half", []))
                        salary_half = work_hours_half * (employee["data_rate"] / 2)
                        salary_full = work_hours * employee["data_rate"]
                        salary = salary_half + salary_full
                        timekeeper_employee = work_hours + (work_hours_half / 2)
                        if employee_id in employee_data1:
                            employee_pre_pay = employee_data1[employee_id]
                            if "data_rate_1" in employee_pre_pay:
                                salary_pre_pay = salary - employee_pre_pay["data_rate_1"]
                            else:
                                salary_pre_pay = salary
                        else:
                            salary_pre_pay = salary

                    else:
                        work_hours = len(employee["timekeeping"])
                        salary = work_hours * employee["data_rate"]
                        timekeeper_employee = work_hours
                        if employee_id in employee_data1:
                            employee_pre_pay = employee_data1[employee_id]
                            if "data_rate_1" in employee_pre_pay:
                                salary_pre_pay = salary - employee_pre_pay["data_rate_1"]
                            else:
                                salary_pre_pay = salary
                        else:
                            salary_pre_pay = salary

                    employee_name.append(f"Employee salary {employee['name']} is: {salary_pre_pay}. ")
                    with open("employee_salary.txt","a+",encoding="utf-8") as a4:
                        a4.write("\n" + employee["name"] + ": " + str(salary))

                    employee_ids.append(employee_id)
                    salary_pre_pays.append(salary_pre_pay) 
                    timekeeper_employees.append(timekeeper_employee)
                    i2 = i2 + 1
                    i3.append(str(i2))

                    message_text = employee_name[i]
                    message = "\n".join([message_text])

                else:
                    employee_name.append(f"Can't see the employee:  {employee_id}. ")

        # Create DataFrame only if the arrays have the same length
        if len(employee_ids) == len(salary_pre_pays):
            data = {
                'STT': i3,
                'Tên': employee_ids,
                'Lương Sau Khi Ứng': salary_pre_pays,
                'Số Công': timekeeper_employees
            }

            df = pd.DataFrame(data)
            existing_data = pd.DataFrame()  # Initialize an empty DataFrame

            try:
                existing_data = pd.read_excel(workbook_path, sheet_name='Lương')
            except FileNotFoundError:
                pass

            df = pd.concat([existing_data, df], ignore_index=True)
            df.to_excel(workbook_path, index=False, sheet_name='Lương')

            # Load the existing Excel file
            workbook = load_workbook(workbook_path)

            # Get the active sheet
            sheet = workbook.active

            # Set the font size and format for the entire columns
            font = Font(size=14, bold=False, name='Times New Roman')
            font1 = Font(size=14, bold=True, name='Times New Roman')

            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

            sheet['A1'].font = font1
            sheet['B1'].font = font1
            sheet['C1'].font = font1
            sheet['D1'].font = font1
            sheet['E1'].font = font1

            column_letter = 'A'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            column_letter = 'B'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            column_letter = 'C'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            column_letter = 'D'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            column_letter = 'E'
            for cell in sheet[column_letter][1:]:
                cell.font = font
                cell.border = thin_border
                column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
                sheet.column_dimensions[column_letter].width = column_width + 10

            workbook.save(workbook_path)
        else:
            print(RED + "Arrays have different lengths. Unable to create DataFrame." + RESET)

        message_box = QMessageBox()
        font = QFont("Times New Roman", 14)  # Specify the font family and size
        message_box.setFont(font)
        message_box.setWindowTitle("Salary")
    
        # Set the HTML-formatted text as the informative text of the QMessageBox
        message_box.setInformativeText(message)

        # Show the QMessageBox
        message_box.exec()

    def export_excel(self):
        # Define list of Excel file names
        if not os.path.exists('cham_cong_ung_thang ' + str(now2) + '.xlsx'):
            file_names = ['cham_cong_luong_thang ' + str(now2) + '.xlsx','cham_cong_thang ' + str(now2) + '.xlsx']  # Change to your actual filename
        else:
            file_names = ['cham_cong_luong_thang ' + str(now2) + '.xlsx', "cham_cong_thang " + str(now2) + ".xlsx", 'cham_cong_ung_thang ' + str(now2) + '.xlsx']  # Change to your actual filename
        # Create a new workbook
        merged_workbook = openpyxl.Workbook()

        # For each Excel file, copy the sheet and merge it into a new workbook
        for file_name in file_names:
            workbook = openpyxl.load_workbook(file_name)  # Load workbook from original Excel file
            for sheet in workbook.sheetnames:
                source_sheet = workbook[sheet]
                target_sheet = merged_workbook.create_sheet(title=sheet)  # Create a new sheet in the new workbook
                for row in source_sheet.iter_rows(values_only=True):
                    target_sheet.append(row)  # Copy data from original sheet to new sheet

        # Delete the default sheet created initially
        default_sheet = merged_workbook['Sheet']
        merged_workbook.remove(default_sheet)

        # Reset the font and format for each sheet in the new workbook
        font = Font(size=14, bold=False, name='Times New Roman')
        font1 = Font(size=14, bold=True, name='Times New Roman')

        for sheet in merged_workbook.sheetnames:
            current_sheet = merged_workbook[sheet]
            
            current_sheet['A1'].font = font1
            current_sheet['B1'].font = font1
            current_sheet['C1'].font = font1
            current_sheet['D1'].font = font1
            
            for column_letter in ['A', 'B', 'C', 'D']:
                for cell in current_sheet[column_letter][1:]:
                    cell.font = font

                column_width = max(len(str(cell.value)) for cell in current_sheet[column_letter])
                current_sheet.column_dimensions[column_letter].width = column_width + 2

        # Save the new workbook to an Excel file
        merged_workbook.save('cham_cong_final_thang 0' + str(int(now2)-1) + " + " + str(now2) + '.xlsx')

        message_text = 'Export file success'

        message_box = QMessageBox()
        font = QFont("Times New Roman", 14)  # Specify the font family and size
        message_box.setFont(font)
        message_box.setWindowTitle("About us")


        # Construct the final message by concatenating the HTML-formatted texts
        message = "\n".join([message_text])

        # Set the HTML-formatted text as the informative text of the QMessageBox
        message_box.setInformativeText(message)

        # Show the QMessageBox
        message_box.exec()
    
    def exit(self):
        exit()

class DeleteWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.delete()

    def delete(self):
        def on_yes_button_click():
            def load_employee_data_backup():
                try:
                    with open("employee_data_backup.txt", "r") as file:
                        employee_data = eval(file.read())
                        return employee_data
                except FileNotFoundError:
                    return {}
                os.remove("employee_data.json")
                if not os.path.exists('employee_data.json'):
                    with open('employee_data.json', 'w') as file:
                        json.dump({}, file)
                # Define the source and destination file names
                source_file_name = 'employee_data_backup.txt'  # Replace with the name of your source file
                destination_file_name = 'employee_data.json'  # Replace with the name of your destination file

                # Open the source file for reading
                try:
                    with open(source_file_name, 'r') as source_file:
                        # Open the destination file for writing
                        with open(destination_file_name, 'w') as destination_file:
                            # Read and copy data from the source file to the destination file
                            data = source_file.read()
                            destination_file.write(data)
                except FileNotFoundError:
                    print(f"File '{source_file_name}' not found.")
                except Exception as e:
                    print(f"An error occurred: {str(e)}")
                else:
                    print("File copied successfully.")
            os.remove("employee_data.txt")
            os.remove("employee_data_half.json")
            os.remove("employee_data_pre_pay.json")
            print(GREEN + "Now you can use it again" + RESET)

        def on_no_button_click():
            QMessageBox.information(window, "Information", "OK! You need to export the file :)")

        choose = self.yes_no_dialog()
        if choose == 'Yes':
            on_yes_button_click()
        else:
            on_no_button_click()

    def yes_no_dialog(self):
        buttons = ("Yes", "No")
        options = []
        # Create the dialog
        dialog = QDialog(self)
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog_layout = QVBoxLayout()

        # Set a custom font for the WindowTitle (DialogTitle)
        title_label = QLabel("Did you export the file?")
        title_label.setFont(QFont("Times New Roman", 16))
        dialog.setWindowTitle(" ")  # Set a space to avoid the default title
        dialog.setWindowFlags(Qt.Dialog | Qt.CustomizeWindowHint)

        dialog_layout.addWidget(title_label)
        dialog_layout.setSizeConstraint(QVBoxLayout.SetFixedSize)
        dialog.setLayout(dialog_layout)

        for button_text in buttons:
            button = QPushButton(button_text)
            button.setFont(QFont("Times New Roman", 14))
            button.clicked.connect(self.dialog_button_clicked)
            options.append(button)

        # Add the buttons to the layout
        for option in options:
            dialog_layout.addWidget(option)

        

        # Execute the dialog and return the selected option
        result = dialog.exec_()
        return self.selected_option

    def dialog_button_clicked(self):
        self.selected_option = self.sender().text()
        dialog = self.sender().parent()
        dialog.close()

    
        
        

if __name__ == '__main__':

    if not os.path.exists('employee_data_pre_pay.json') and not os.path.exists("employee_data_half.json"):
        with open('employee_data_pre_pay.json', 'w') as file:
            json.dump({}, file)
        with open('employee_data_half.json', 'w') as file:
            json.dump({}, file)

    app = QApplication(sys.argv)
    window = WelcomeWindow()
    window.show()

    if now1 == '01':
        window1 = DeleteWindow()
        timer = QTimer()
        timer.timeout.connect(window1.show)
        timer.start(5)  # 5 seconds delay
    else:
        time.sleep(5)
    
    ex = MenuExample()
    sys.exit(app.exec_())
